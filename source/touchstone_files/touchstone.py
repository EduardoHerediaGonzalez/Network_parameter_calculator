import os.path

from openpyxl.reader.excel import load_workbook

import source.config_parameters as cfg
from source.network_parameter_conversions import *

touchstone_header_info = list()
hashtag_symbol_row = 0

def get_frequency_parameters_from_touchstone_file(excel_touchstone_sheet):
    row_counter = 0
    symbol = ''
    global touchstone_header_info
    global hashtag_symbol_row

    while symbol != '#':

        row_counter = row_counter + 1
        symbol = excel_touchstone_sheet.cell(row=row_counter, column=cfg.EXCEL_COLUMN_A).value
        symbol = symbol.split()
        symbol = symbol[0]

    hashtag_symbol_row = row_counter

    touchstone_header_info = excel_touchstone_sheet.cell(row=row_counter, column= cfg.EXCEL_COLUMN_A).value.split()

    start_frequency_row = row_counter + 1
    start_frequency = excel_touchstone_sheet.cell(row=start_frequency_row, column= cfg.EXCEL_COLUMN_A).value
    start_frequency = start_frequency.split()
    start_frequency = start_frequency[0]
    start_frequency = float(start_frequency)

    next_frequency_row = start_frequency_row + 1
    next_frequency = excel_touchstone_sheet.cell(row=next_frequency_row, column= cfg.EXCEL_COLUMN_A).value
    next_frequency = next_frequency.split()
    next_frequency = next_frequency[0]
    next_frequency = float(next_frequency)

    analysis_frequency_step = next_frequency - start_frequency

    end_frequency_row = excel_touchstone_sheet.max_row
    end_frequency = excel_touchstone_sheet.cell(row=end_frequency_row, column= cfg.EXCEL_COLUMN_A).value
    end_frequency = end_frequency.split()
    end_frequency = end_frequency[0]
    end_frequency = float(end_frequency)

    return start_frequency, end_frequency, analysis_frequency_step

def get_touchstone_parameters(excel_touchstone_sheet, at_frequency):
    row_counter = 0
    frequency = ''
    touchstone_parameters = list()

    while frequency != str(int(at_frequency)):
        row_counter = row_counter + 1

        frequency = excel_touchstone_sheet.cell(row=row_counter, column=cfg.EXCEL_COLUMN_A).value
        frequency = frequency.split()
        frequency = frequency[0]

    parameters = excel_touchstone_sheet.cell(row=row_counter, column=cfg.EXCEL_COLUMN_A).value
    parameters = parameters.split()
    touchstone_parameters.append(parameters[1])
    touchstone_parameters.append(parameters[2])
    touchstone_parameters.append(parameters[3])
    touchstone_parameters.append(parameters[4])
    touchstone_parameters.append(parameters[5])
    touchstone_parameters.append(parameters[6])
    touchstone_parameters.append(parameters[7])
    touchstone_parameters.append(parameters[8])

    return touchstone_parameters

def write_touchstone_file(network_parameters_workbook):
    excel_touchstone_workbook = load_workbook(filename=os.path.join(os.getcwd(), cfg.FOLDER_TOUCHSTONE_FILES, cfg.EXCEL_TOUCHSTONE_FILE_NAME))
    excel_touchstone_sheet = excel_touchstone_workbook[cfg.EXCEL_DEFAULT_SHEET_NAME]
    excel_s_parameters_sheet = network_parameters_workbook[cfg.EXCEL_S_PARAMETERS_SHEET]

    excel_touchstone_row = hashtag_symbol_row + 1
    total_of_rows = excel_s_parameters_sheet.max_row

    for row in range(cfg.EXCEL_INITIAL_ROW, total_of_rows + 1):
        frequency = str(int(excel_s_parameters_sheet.cell(row=row, column=cfg.EXCEL_COLUMN_A).value))
        parameter_s11 = complex(excel_s_parameters_sheet.cell(row=row, column=cfg.EXCEL_COLUMN_B).value)
        parameter_s12 = complex(excel_s_parameters_sheet.cell(row=row, column=cfg.EXCEL_COLUMN_C).value)
        parameter_s21 = complex(excel_s_parameters_sheet.cell(row=row, column=cfg.EXCEL_COLUMN_D).value)
        parameter_s22 = complex(excel_s_parameters_sheet.cell(row=row, column=cfg.EXCEL_COLUMN_E).value)


        touchstone_row = (frequency + ' ' + str(parameter_s11.real) + ' ' + str(parameter_s11.imag) + ' ' +
                          str(parameter_s21.real) + ' ' + str(parameter_s21.imag) + ' ' +
                          str(parameter_s12.real) + ' ' + str(parameter_s12.imag) + ' ' +
                          str(parameter_s22.real) + ' ' + str(parameter_s22.imag))

        excel_touchstone_sheet.cell(row=excel_touchstone_row, column=cfg.EXCEL_COLUMN_A).value = touchstone_row
        excel_touchstone_row = excel_touchstone_row + 1

    excel_touchstone_workbook.save(os.path.join(os.getcwd(), cfg.FOLDER_TOUCHSTONE_FILES, cfg.EXCEL_TOUCHSTONE_FILE_NAME))

# Definition of the class that represents a Touchstone file
class TouchstoneFile:
    # Private class attributes
    __excel_touchstone_workbook = 0
    __excel_touchstone_sheet = 0
    __touchstone_parameters: list
    __matrix_abcd: matrix

    # Class constructors
    def __init__(self, excel_touchstone_workbook):
        self.__excel_touchstone_workbook = excel_touchstone_workbook
        self.__excel_touchstone_sheet = self.__excel_touchstone_workbook[cfg.EXCEL_DEFAULT_SHEET_NAME]

    # Public class methods
    def get_ABCD_matrix(self, at_frequency):
        self.__touchstone_parameters = get_touchstone_parameters(excel_touchstone_sheet=self.__excel_touchstone_sheet, at_frequency=at_frequency)
        parameter_s11 = complex(float(self.__touchstone_parameters[0]), float(self.__touchstone_parameters[1]))
        parameter_s12 = complex(float(self.__touchstone_parameters[4]), float(self.__touchstone_parameters[5]))
        parameter_s21 = complex(float(self.__touchstone_parameters[2]), float(self.__touchstone_parameters[3]))
        parameter_s22 = complex(float(self.__touchstone_parameters[6]), float(self.__touchstone_parameters[7]))
        s_matrix = np.matrix([[parameter_s11, parameter_s12],[parameter_s21, parameter_s22]])

        self.__matrix_abcd = convert_S_matrix_to_ABCD_matrix(s_matrix=s_matrix, z_0=float(touchstone_header_info[cfg.TOUCHSTONE_CHARACTERISTIC_IMPEDANCE_INDEX]))

        return self.__matrix_abcd
