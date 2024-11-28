import os
import shutil
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk

import pandas as pd
from openpyxl import *
from pandas.io.sas.sas_constants import row_count_offset_multiplier

from source.common_two_port_circuits import *
from source.plot_parameters import *
from source.touchstone_files.touchstone import *

excel_base_template_workbook = load_workbook(filename=os.path.join(os.getcwd(), cfg.FOLDER_EXCEL_FILES, cfg.EXCEL_BASE_TEMPLATE_FILE))
excel_base_template_workbook.save(os.path.join(os.getcwd(), cfg.FOLDER_EXCEL_FILES, cfg.EXCEL_NETWORK_PARAMETERS_FILE))
excel_base_template_workbook.close()
del excel_base_template_workbook

excel_network_parameters_workbook = load_workbook(filename=os.path.join(os.getcwd(), cfg.FOLDER_EXCEL_FILES, cfg.EXCEL_NETWORK_PARAMETERS_FILE))
excel_network_info_sheet = excel_network_parameters_workbook[cfg.EXCEL_NETWORK_INFO_SHEET]

excel_circuit_counter_row = cfg.EXCEL_INITIAL_ROW_NETWORK_ID
sub_networks = list()
sub_networks_interconnection = list()
frequency_range = list()
touchstone_file_counter = 0
parameters_to_calculate = ''

def add_touchstone_file():
    touchstone_file_path = touchstone_file_entry.get()
    global touchstone_file_counter
    global counter_of_sub_networks
    global excel_circuit_counter_row

    if os.path.exists(touchstone_file_path):
        shutil.copy2(touchstone_file_path, os.path.join(os.getcwd(), cfg.FOLDER_TOUCHSTONE_FILES, cfg.TXT_TOUCHSTONE_FILE_NAME))

        data_frame = pd.read_csv(os.path.join(os.getcwd(), cfg.FOLDER_TOUCHSTONE_FILES, cfg.TXT_TOUCHSTONE_FILE_NAME), sep=',')
        data_frame.to_excel(os.path.join(os.getcwd(), cfg.FOLDER_TOUCHSTONE_FILES, cfg.EXCEL_TOUCHSTONE_FILE_NAME), index=False)
        os.remove(os.path.join(os.getcwd(), cfg.FOLDER_TOUCHSTONE_FILES, cfg.TXT_TOUCHSTONE_FILE_NAME))

        sub_networks_counter = counter_of_sub_networks.get()
        sub_networks_counter = sub_networks_counter + 1

        excel_network_info_sheet.cell(row=excel_circuit_counter_row, column=cfg.EXCEL_COLUMN_A, value=str(sub_networks_counter))
        excel_network_info_sheet.cell(row=excel_circuit_counter_row, column=cfg.EXCEL_COLUMN_B, value=cfg.TOUCHSTONE_CONNECTION)
        excel_network_info_sheet.cell(row=excel_circuit_counter_row, column=cfg.EXCEL_COLUMN_C, value=cfg.EXCEL_TOUCHSTONE_FILE_NAME)
        excel_network_parameters_workbook.save(os.path.join(os.getcwd(), cfg.FOLDER_EXCEL_FILES, cfg.EXCEL_NETWORK_PARAMETERS_FILE))
        excel_touchstone_workbook = load_workbook(filename=os.path.join(os.getcwd(), cfg.FOLDER_TOUCHSTONE_FILES, cfg.EXCEL_TOUCHSTONE_FILE_NAME))
        excel_touchstone_sheet = excel_touchstone_workbook[cfg.EXCEL_DEFAULT_SHEET_NAME]

        (start_frequency_value, end_frequency_value, analysis_frequency_step_value) = get_frequency_parameters_from_touchstone_file(excel_touchstone_sheet)

        start_frequency = get_frequency_with_prefixed(start_frequency_value)
        end_frequency = get_frequency_with_prefixed(end_frequency_value)
        analysis_frequency_step = get_frequency_with_prefixed(analysis_frequency_step_value)

        start_frequency = start_frequency.split(' ')
        start_frequency_prefix = start_frequency[1]
        start_frequency = start_frequency[0]
        reset_start_frequency_entry.set(start_frequency)
        start_frequency_combobox.current(cfg.FREQUENCY_PREFIXES.index(start_frequency_prefix))

        end_frequency = end_frequency.split(' ')
        end_frequency_prefix = end_frequency[1]
        end_frequency = end_frequency[0]
        reset_end_frequency_entry.set(end_frequency)
        end_frequency_combobox.current(cfg.FREQUENCY_PREFIXES.index(end_frequency_prefix))

        analysis_frequency_step = analysis_frequency_step.split(' ')
        analysis_frequency_step_prefix = analysis_frequency_step[1]
        analysis_frequency_step = analysis_frequency_step[0]
        reset_analysis_frequency_step_entry.set(analysis_frequency_step)
        analysis_frequency_step_combobox.current(cfg.FREQUENCY_PREFIXES.index(analysis_frequency_step_prefix))

        reset_touchstone_file_entry.set('')
        counter_of_sub_networks.set(sub_networks_counter)
        touchstone_file_counter = 1

        if sub_networks_counter == 1:
            save_sub_networks_button.config(state='normal')

    else:
        messagebox.showerror(title='Error', message='Touchstone file not found. Check that the path or filename is correct.')

def add_sub_network():
    type_of_circuit = type_of_circuit_combobox.get()
    type_of_interconnection = type_of_interconnection_combobox.get()
    element_a = element_A_combobox.get()
    element_a_value = element_A_entry.get()
    element_b = element_B_combobox.get()
    element_b_value = element_B_entry.get()
    element_c = element_C_combobox.get()
    element_c_value = element_C_entry.get()


    if type_of_circuit == '':
        messagebox.showerror(title='Error', message='Type of circuit is empty')

    elif element_a == '' and element_b == '' and element_c == '':
        messagebox.showerror(title='Error', message='No element selected')

    elif element_a_value == '' and element_b_value == '' and element_c_value == '':
        messagebox.showerror(title='Error', message='No element value')

    else:
        global excel_circuit_counter_row

        sub_networks_counter = counter_of_sub_networks.get()
        sub_networks_counter = sub_networks_counter + 1

        if type_of_interconnection == '':
            type_of_interconnection = cfg.SUB_NETWORK_DEFAULT_CONNECTION

        excel_network_info_sheet.cell(row=excel_circuit_counter_row, column=cfg.EXCEL_COLUMN_A, value=str(sub_networks_counter))
        excel_network_info_sheet.cell(row = excel_circuit_counter_row, column = cfg.EXCEL_COLUMN_B, value = type_of_interconnection)
        excel_network_info_sheet.cell(row = excel_circuit_counter_row, column = cfg.EXCEL_COLUMN_C, value = type_of_circuit)
        excel_network_info_sheet.cell(row = excel_circuit_counter_row, column = cfg.EXCEL_COLUMN_D, value = element_a)
        excel_network_info_sheet.cell(row = excel_circuit_counter_row, column = cfg.EXCEL_COLUMN_E, value = join_value_and_prefix(element_a_value, element_a))
        excel_network_info_sheet.cell(row = excel_circuit_counter_row, column = cfg.EXCEL_COLUMN_F, value = element_b)
        excel_network_info_sheet.cell(row = excel_circuit_counter_row, column = cfg.EXCEL_COLUMN_G, value = join_value_and_prefix(element_b_value, element_b))
        excel_network_info_sheet.cell(row = excel_circuit_counter_row, column = cfg.EXCEL_COLUMN_H, value = element_c)
        excel_network_info_sheet.cell(row = excel_circuit_counter_row, column = cfg.EXCEL_COLUMN_I, value = join_value_and_prefix(element_c_value, element_c))
        excel_network_parameters_workbook.save(os.path.join(os.getcwd(), cfg.FOLDER_EXCEL_FILES, cfg.EXCEL_NETWORK_PARAMETERS_FILE))

        excel_circuit_counter_row = excel_circuit_counter_row + 1

        counter_of_sub_networks.set(sub_networks_counter)
        reset_type_of_circuit_combobox.set('')
        reset_type_of_connection_combobox.set('')
        reset_element_A_combobox.set('')
        reset_element_A_entry.set('')
        reset_element_B_combobox.set('')
        reset_element_B_entry.set('')
        reset_element_C_combobox.set('')
        reset_element_C_entry.set('')

        if sub_networks_counter == 1:
            save_sub_networks_button.config(state='normal')

def save_sub_networks():
    type_of_circuit_combobox.config(state='disable')
    type_of_interconnection_combobox.config(state='disable')
    element_A_combobox.config(state='disable')
    element_A_entry.config(state='disable')
    element_B_combobox.config(state='disable')
    element_B_entry.config(state='disable')
    element_C_combobox.config(state='disable')
    element_C_entry.config(state='disable')
    touchstone_file_button.config(state='disable')
    touchstone_file_entry.config(state='disable')
    add_sub_network_button.config(state='disable')
    save_sub_networks_button.config(state='disable')

    if touchstone_file_counter == 0:
        start_frequency_entry.config(state='normal')
        start_frequency_combobox.config(state='normal')
        end_frequency_entry.config(state='normal')
        end_frequency_combobox.config(state='normal')
        analysis_frequency_step_entry.config(state='normal')
        analysis_frequency_step_combobox.config(state='normal')

    save_frequency_parameters_button.config(state='normal')

def save_frequency_parameters():
    start_frequency = start_frequency_entry.get()
    start_frequency_prefix = start_frequency_combobox.get()

    end_frequency = end_frequency_entry.get()
    end_frequency_prefix = end_frequency_combobox.get()

    analysis_frequency_step = analysis_frequency_step_entry.get()
    analysis_frequency_step_prefix = analysis_frequency_step_combobox.get()

    if start_frequency == '' or start_frequency_prefix == '' or end_frequency == '' or end_frequency_prefix == '' or analysis_frequency_step == '':
        messagebox.showerror(title='Error', message='One or more fields are empty')
        
    else:
        start_frequency_value = float(start_frequency) * cfg.FREQUENCY_PREFIXES_TO_VALUE[start_frequency_prefix]
        end_frequency_value = float(end_frequency) * cfg.FREQUENCY_PREFIXES_TO_VALUE[end_frequency_prefix]
        analysis_frequency_step_value = float(analysis_frequency_step) * cfg.FREQUENCY_PREFIXES_TO_VALUE[analysis_frequency_step_prefix]

        total_of_frequency_points =  round(((end_frequency_value - start_frequency_value) / analysis_frequency_step_value) + 1)

        frequency_range.append(start_frequency_value)
        current_frequency_point_value = start_frequency_value

        for frequency_point_counter in range(1, total_of_frequency_points):
            current_frequency_point_value = current_frequency_point_value + analysis_frequency_step_value
            frequency_range.append(current_frequency_point_value)

        excel_network_info_sheet.cell(row=cfg.EXCEL_INITIAL_ROW, column=cfg.EXCEL_COLUMN_B).value = start_frequency + ' ' + start_frequency_prefix
        excel_network_info_sheet.cell(row=cfg.EXCEL_INITIAL_ROW + 1, column=cfg.EXCEL_COLUMN_B).value = end_frequency + ' ' + end_frequency_prefix
        excel_network_info_sheet.cell(row =cfg.EXCEL_INITIAL_ROW + 2, column = cfg.EXCEL_COLUMN_B).value = analysis_frequency_step + ' ' + analysis_frequency_step_prefix
        excel_network_parameters_workbook.save(os.path.join(os.getcwd(), cfg.FOLDER_EXCEL_FILES, cfg.EXCEL_NETWORK_PARAMETERS_FILE))

        start_frequency_entry.config(state='disable')
        start_frequency_combobox.config(state='disable')
        end_frequency_entry.config(state='disable')
        end_frequency_combobox.config(state='disable')
        analysis_frequency_step_entry.config(state='disable')
        save_frequency_parameters_button.config(state='disable')
        calculate_parameters_button.config(state='normal')
        parameters_to_calculate_combobox.config(state='normal')

def calculate_parameters():
    global sub_networks
    global sub_networks_interconnection
    global parameters_to_calculate

    excel_ABCD_parameters_sheet = excel_network_parameters_workbook[cfg.EXCEL_ABCD_PARAMETERS_SHEET]
    excel_Z_parameters_sheet = excel_network_parameters_workbook[cfg.EXCEL_Z_PARAMETERS_SHEET]
    excel_Y_parameters_sheet = excel_network_parameters_workbook[cfg.EXCEL_Y_PARAMETERS_SHEET]
    excel_S_parameters_sheet = excel_network_parameters_workbook[cfg.EXCEL_S_PARAMETERS_SHEET]

    parameters_to_calculate = parameters_to_calculate_combobox.get()

    if parameters_to_calculate != '':
        excel_network_info_sheet.cell(row=cfg.EXCEL_INITIAL_ROW + 4, column=cfg.EXCEL_COLUMN_B).value = cfg.DEFAULT_NETWORK_TYPE
        excel_network_info_sheet.cell(row=cfg.EXCEL_INITIAL_ROW + 5, column=cfg.EXCEL_COLUMN_B).value = parameters_to_calculate
        excel_network_parameters_workbook.save(os.path.join(os.getcwd(), cfg.FOLDER_EXCEL_FILES, cfg.EXCEL_NETWORK_PARAMETERS_FILE))

        get_sub_networks_info()

        row_counter = cfg.EXCEL_INITIAL_ROW

        for frequency in frequency_range:
            total_ABCD_matrix = get_total_ABCD_matrix(at_frequency=frequency)
            total_Z_matrix = convert_ABCD_matrix_to_Z_matrix(abcd_matrix=total_ABCD_matrix)
            total_Y_matrix = convert_ABCD_matrix_to_Y_matrix(abcd_matrix=total_ABCD_matrix)
            total_S_matrix = convert_ABCD_matrix_to_S_matrix(abcd_matrix=total_ABCD_matrix, z_0=50)

            (parameter_a, parameter_b , parameter_c, parameter_d, delta) = get_parameters_and_delta_from_matrix(total_ABCD_matrix)
            (parameter_z11, parameter_z12, parameter_z21, parameter_z22, delta) = get_parameters_and_delta_from_matrix(total_Z_matrix)
            (parameter_y11, parameter_y12, parameter_y21, parameter_y22, delta) = get_parameters_and_delta_from_matrix(total_Y_matrix)
            (parameter_s11, parameter_s12, parameter_s21, parameter_s22, delta) = get_parameters_and_delta_from_matrix(total_S_matrix)

            excel_ABCD_parameters_sheet.cell(row=row_counter, column=cfg.EXCEL_COLUMN_A).value = frequency
            excel_ABCD_parameters_sheet.cell(row=row_counter, column=cfg.EXCEL_COLUMN_B).value = str(parameter_a).strip('()')
            excel_ABCD_parameters_sheet.cell(row=row_counter, column=cfg.EXCEL_COLUMN_C).value = str(parameter_b).strip('()')
            excel_ABCD_parameters_sheet.cell(row=row_counter, column=cfg.EXCEL_COLUMN_D).value = str(parameter_c).strip('()')
            excel_ABCD_parameters_sheet.cell(row=row_counter, column=cfg.EXCEL_COLUMN_E).value = str(parameter_d).strip('()')

            excel_Z_parameters_sheet.cell(row=row_counter, column=cfg.EXCEL_COLUMN_A).value = frequency
            excel_Z_parameters_sheet.cell(row=row_counter, column=cfg.EXCEL_COLUMN_B).value = str(parameter_z11).strip('()')
            excel_Z_parameters_sheet.cell(row=row_counter, column=cfg.EXCEL_COLUMN_C).value = str(parameter_z12).strip('()')
            excel_Z_parameters_sheet.cell(row=row_counter, column=cfg.EXCEL_COLUMN_D).value = str(parameter_z21).strip('()')
            excel_Z_parameters_sheet.cell(row=row_counter, column=cfg.EXCEL_COLUMN_E).value = str(parameter_z22).strip('()')

            excel_Y_parameters_sheet.cell(row=row_counter, column=cfg.EXCEL_COLUMN_A).value = frequency
            excel_Y_parameters_sheet.cell(row=row_counter, column=cfg.EXCEL_COLUMN_B).value = str(parameter_y11).strip('()')
            excel_Y_parameters_sheet.cell(row=row_counter, column=cfg.EXCEL_COLUMN_C).value = str(parameter_y12).strip('()')
            excel_Y_parameters_sheet.cell(row=row_counter, column=cfg.EXCEL_COLUMN_D).value = str(parameter_y21).strip('()')
            excel_Y_parameters_sheet.cell(row=row_counter, column=cfg.EXCEL_COLUMN_E).value = str(parameter_y22).strip('()')

            excel_S_parameters_sheet.cell(row=row_counter, column=cfg.EXCEL_COLUMN_A).value = frequency
            excel_S_parameters_sheet.cell(row=row_counter, column=cfg.EXCEL_COLUMN_B).value = str(parameter_s11).strip('()')
            excel_S_parameters_sheet.cell(row=row_counter, column=cfg.EXCEL_COLUMN_C).value = str(parameter_s12).strip('()')
            excel_S_parameters_sheet.cell(row=row_counter, column=cfg.EXCEL_COLUMN_D).value = str(parameter_s21).strip('()')
            excel_S_parameters_sheet.cell(row=row_counter, column=cfg.EXCEL_COLUMN_E).value = str(parameter_s22).strip('()')

            excel_network_parameters_workbook.save(os.path.join(os.getcwd(), cfg.FOLDER_EXCEL_FILES, cfg.EXCEL_NETWORK_PARAMETERS_FILE))

            row_counter = row_counter + 1

        plot_parameters_button.config(state='normal')
        plot_parameters_in_format_combobox.config(state='normal')

        messagebox.showinfo(title='Info', message='Parameters calculated')

    else:
        messagebox.showerror(title='Error', message='Empty parameters to calculate')

def plot_parameters():
    parameter_a = list()
    parameter_b = list()
    parameter_c = list()
    parameter_d = list()

    if parameters_to_calculate == cfg.NETWORK_PARAMETERS[0]:
        excel_parameters_sheet = excel_network_parameters_workbook[cfg.EXCEL_Z_PARAMETERS_SHEET]

    elif parameters_to_calculate == cfg.NETWORK_PARAMETERS[1]:
        excel_parameters_sheet = excel_network_parameters_workbook[cfg.EXCEL_Y_PARAMETERS_SHEET]

    elif parameters_to_calculate == cfg.NETWORK_PARAMETERS[2]:
        excel_parameters_sheet = excel_network_parameters_workbook[cfg.EXCEL_ABCD_PARAMETERS_SHEET]

    else:
        excel_parameters_sheet = excel_network_parameters_workbook[cfg.EXCEL_S_PARAMETERS_SHEET]

    for row_counter in range(cfg.EXCEL_INITIAL_ROW, excel_parameters_sheet.max_row + 1):

        a = excel_parameters_sheet.cell(row=row_counter, column=cfg.EXCEL_COLUMN_B).value
        b = excel_parameters_sheet.cell(row=row_counter, column=cfg.EXCEL_COLUMN_C).value
        c = excel_parameters_sheet.cell(row=row_counter, column=cfg.EXCEL_COLUMN_D).value
        d = excel_parameters_sheet.cell(row=row_counter, column=cfg.EXCEL_COLUMN_E).value

        parameter_a.append(complex(a))
        parameter_b.append(complex(b))
        parameter_c.append(complex(c))
        parameter_d.append(complex(d))

    format_to_plot = plot_parameters_in_format_combobox.get()

    if format_to_plot == cfg.PLOT_FORMATS[0]:
        plot_magnitude_vs_frequency(frequency_range, parameter_a, parameter_b, parameter_c, parameter_d)

    elif format_to_plot == cfg.PLOT_FORMATS[1]:
        plot_phase_vs_frequency(frequency_range, parameter_a, parameter_b, parameter_c, parameter_d)

    elif format_to_plot == cfg.PLOT_FORMATS[2]:
        plot_r_i_vs_frequency(frequency_range, parameter_a, parameter_b, parameter_c, parameter_d)

    elif format_to_plot == cfg.PLOT_FORMATS[3]:
        plot_polar(parameter_a, parameter_b, parameter_c, parameter_d)

    elif format_to_plot == cfg.PLOT_FORMATS[4]:
        plot_smith_chart(frequency_range, parameter_a, parameter_b, parameter_c, parameter_d)

def get_frequency_with_prefixed(frequency_value: float):
    frequency_value__length = len(str(int(frequency_value)))
    frequency_with_prefixed = ''

    if frequency_value__length <= cfg.THREE_DIGITS:
        frequency_value = frequency_value / cfg.FREQUENCY_PREFIXES_TO_VALUE[cfg.FREQUENCY_PREFIXES[0]]
        frequency_with_prefixed = str(round(frequency_value,2)) + ' ' + cfg.FREQUENCY_PREFIXES[0]

    elif frequency_value__length <= cfg.SIX_DIGITS:
        frequency_value = frequency_value / cfg.FREQUENCY_PREFIXES_TO_VALUE[cfg.FREQUENCY_PREFIXES[1]]
        frequency_with_prefixed = str(round(frequency_value,2)) + ' ' + cfg.FREQUENCY_PREFIXES[1]

    elif frequency_value__length <= cfg.NINE_DIGITS:
        frequency_value = frequency_value / cfg.FREQUENCY_PREFIXES_TO_VALUE[cfg.FREQUENCY_PREFIXES[2]]
        frequency_with_prefixed = str(round(frequency_value,2)) + ' ' + cfg.FREQUENCY_PREFIXES[2]

    elif frequency_value__length <= cfg.TWELVE_DIGITS:
        frequency_value = frequency_value / cfg.FREQUENCY_PREFIXES_TO_VALUE['GHz']
        frequency_with_prefixed = str(round(frequency_value,2)) + ' ' + cfg.FREQUENCY_PREFIXES[3]

    return  frequency_with_prefixed

def convert_string_to_value(string_value):
    num_str = ''
    prefix_str = ''
    _value = 0

    for character_value in string_value:
        if character_value.isdigit() or character_value == '.':
            num_str = num_str + character_value
        else:
            prefix_str = prefix_str + character_value

    if len(prefix_str) != 0:
        prefix_str = prefix_str[0]
        prefix_value = cfg.MAGNITUDES_PREFIXES_TO_VALUE[prefix_str]
        num_value = float(num_str)
        _value = num_value * prefix_value

    else:
        _value = float(string_value)

    return _value

def join_value_and_prefix(value: str, element_type: str):
    num_str = ''
    prefix_str = ''

    for character_value in value:
        if character_value.isdigit() or character_value == '.':
            num_str = num_str + character_value
        else:
            if character_value != ' ':
                if element_type == cfg.ELEMENT_TYPES[0] and character_value == 'm':
                    character_value = character_value.title()

                if character_value == 'f' or character_value == 'h' or character_value == 'k':
                    character_value = character_value.title()

                if character_value == 'N' or character_value == 'P' or character_value == 'U':
                    character_value = character_value.lower()

                prefix_str = prefix_str + character_value


    value = num_str + prefix_str

    return value

def get_sub_networks_info():
    sub_network_id_counter = cfg.EXCEL_INITIAL_ROW_NETWORK_ID
    global sub_networks
    global sub_networks_interconnection

    while excel_network_info_sheet.cell(row=sub_network_id_counter, column=cfg.EXCEL_COLUMN_A).value is not None:
        interconnection_type = excel_network_info_sheet.cell(row=sub_network_id_counter, column=cfg.EXCEL_COLUMN_B).value
        sub_networks_interconnection.append(interconnection_type)

        circuit_type = excel_network_info_sheet.cell(row=sub_network_id_counter, column=cfg.EXCEL_COLUMN_C).value

        if circuit_type == 'Series impedance':
            element = excel_network_info_sheet.cell(row=sub_network_id_counter, column=cfg.EXCEL_COLUMN_D).value
            element_value = excel_network_info_sheet.cell(row=sub_network_id_counter, column=cfg.EXCEL_COLUMN_E).value

            sub_network = SeriesImpedanceCircuit(type_of_element=element, element_value=convert_string_to_value(element_value))
            sub_networks.append(sub_network)

        elif circuit_type == 'Shunt impedance':
            element = excel_network_info_sheet.cell(row=sub_network_id_counter, column=cfg.EXCEL_COLUMN_D).value
            element_value = excel_network_info_sheet.cell(row=sub_network_id_counter, column=cfg.EXCEL_COLUMN_E).value

            sub_network = ShuntImpedanceCircuit(type_of_element=element, element_value=convert_string_to_value(element_value))
            sub_networks.append(sub_network)

        elif circuit_type == 'T':
            element_a = excel_network_info_sheet.cell(row=sub_network_id_counter, column=cfg.EXCEL_COLUMN_D).value
            element_a_value = excel_network_info_sheet.cell(row=sub_network_id_counter, column=cfg.EXCEL_COLUMN_E).value
            element_b = excel_network_info_sheet.cell(row=sub_network_id_counter, column=cfg.EXCEL_COLUMN_F).value
            element_b_value = excel_network_info_sheet.cell(row=sub_network_id_counter, column=cfg.EXCEL_COLUMN_G).value
            element_c = excel_network_info_sheet.cell(row=sub_network_id_counter, column=cfg.EXCEL_COLUMN_H).value
            element_c_value = excel_network_info_sheet.cell(row=sub_network_id_counter, column=cfg.EXCEL_COLUMN_I).value

            sub_network = TCircuit(type_of_element_a=element_a, element_a_value=convert_string_to_value(element_a_value),
                                   type_of_element_b=element_b, element_b_value=convert_string_to_value(element_b_value),
                                   type_of_element_c=element_c, element_c_value=convert_string_to_value(element_c_value))
            sub_networks.append(sub_network)

        elif circuit_type == 'Pi':
            element_a = excel_network_info_sheet.cell(row=sub_network_id_counter, column=cfg.EXCEL_COLUMN_D).value
            element_a_value = excel_network_info_sheet.cell(row=sub_network_id_counter, column=cfg.EXCEL_COLUMN_E).value
            element_b = excel_network_info_sheet.cell(row=sub_network_id_counter, column=cfg.EXCEL_COLUMN_F).value
            element_b_value = excel_network_info_sheet.cell(row=sub_network_id_counter, column=cfg.EXCEL_COLUMN_G).value
            element_c = excel_network_info_sheet.cell(row=sub_network_id_counter, column=cfg.EXCEL_COLUMN_H).value
            element_c_value = excel_network_info_sheet.cell(row=sub_network_id_counter, column=cfg.EXCEL_COLUMN_I).value

            sub_network = PiCircuit(type_of_element_a=element_a, element_a_value=convert_string_to_value(element_a_value),
                                   type_of_element_b=element_b, element_b_value=convert_string_to_value(element_b_value),
                                   type_of_element_c=element_c, element_c_value=convert_string_to_value(element_c_value))
            sub_networks.append(sub_network)

        sub_network_id_counter = sub_network_id_counter + 1

    # return sub_networks, sub_networks_interconnection

def get_total_ABCD_matrix(at_frequency: float):
    global sub_networks
    global sub_networks_interconnection

    total_ABCD_matrix = np.matrix([[1, 0],[0, 1]], dtype=complex)
    total_Z_matrix = np.matrix([[0, 0],[0, 0]],dtype=complex)
    total_Y_matrix = np.matrix([[0, 0], [0, 0]], dtype=complex)
    _sub_network_ABCD_matrix = np.matrix([[0, 0], [0, 0]], dtype=complex)
    sub_network_index = 0
    total_of_sub_networks = len(sub_networks)

    for sub_network_interconnection in sub_networks_interconnection:
        _sub_network_ABCD_matrix = sub_networks[sub_network_index].get_ABCD_matrix(at_frequency=at_frequency)

        if sub_network_interconnection == cfg.SUB_NETWORK_DEFAULT_CONNECTION:
            total_ABCD_matrix = total_ABCD_matrix * _sub_network_ABCD_matrix

        elif sub_network_interconnection == cfg.INTERCONNECTION_TYPES[cfg.SERIES_CONNECTION_INDEX]:
            total_Z_matrix = total_Z_matrix + convert_ABCD_matrix_to_Z_matrix(_sub_network_ABCD_matrix)

            if sub_network_index < (total_of_sub_networks - 1):
                if sub_networks_interconnection[sub_network_index + 1] != cfg.INTERCONNECTION_TYPES[cfg.SERIES_CONNECTION_INDEX]:
                    total_ABCD_matrix = total_ABCD_matrix * convert_Z_matrix_to_ABCD_matrix(total_Z_matrix)

            else:
                total_ABCD_matrix = total_ABCD_matrix * convert_Z_matrix_to_ABCD_matrix(total_Z_matrix)

        elif sub_network_interconnection == cfg.INTERCONNECTION_TYPES[cfg.PARALLEL_CONNECTION_INDEX]:
            total_Y_matrix = total_Y_matrix + convert_ABCD_matrix_to_Y_matrix(_sub_network_ABCD_matrix)

            if sub_network_index < (total_of_sub_networks - 1):
                if sub_networks_interconnection[sub_network_index + 1] != cfg.INTERCONNECTION_TYPES[cfg.PARALLEL_CONNECTION_INDEX]:
                    total_ABCD_matrix = total_ABCD_matrix * convert_Y_matrix_to_ABCD_matrix(total_Y_matrix)

            else:
                total_ABCD_matrix = total_ABCD_matrix * convert_Y_matrix_to_ABCD_matrix(total_Y_matrix)

        sub_network_index = sub_network_index + 1

    return total_ABCD_matrix

mainWindow = tk.Tk()
mainWindow.title("Two-port network parameter calculator")
mainWindow.geometry('800x600')

counter_of_sub_networks = tk.IntVar(mainWindow, 0)
reset_type_of_circuit_combobox = tk.StringVar(mainWindow, '')
reset_type_of_connection_combobox = tk.StringVar(mainWindow, '')
reset_element_A_combobox = tk.StringVar(mainWindow, '')
reset_element_A_entry = tk.StringVar(mainWindow, '')
reset_element_B_combobox = tk.StringVar(mainWindow, '')
reset_element_B_entry = tk.StringVar(mainWindow, '')
reset_element_C_combobox = tk.StringVar(mainWindow, '')
reset_element_C_entry = tk.StringVar(mainWindow, '')
reset_touchstone_file_entry = tk.StringVar(mainWindow, '')
reset_start_frequency_entry = tk.StringVar(mainWindow, '')
reset_end_frequency_entry = tk.StringVar(mainWindow, '')
reset_analysis_frequency_step_entry = tk.StringVar(mainWindow, '')

# Start of frame 1 #
row = 0

frame_1 = ttk.Frame(mainWindow)
frame_1.grid(row=row, column=0)

spacer_1 = ttk.Label(frame_1)
spacer_2 = ttk.Label(frame_1)
spacer_3 = ttk.Label(frame_1)
spacer_4 = ttk.Label(frame_1)

type_of_circuit_label = ttk.Label(frame_1, text='Type of circuit')
type_of_circuit_label.grid(row=row, column=0)
type_of_circuit_combobox = ttk.Combobox(frame_1, values=cfg.CIRCUIT_TYPES, state='normal', textvariable=reset_type_of_circuit_combobox)
type_of_circuit_combobox.grid(row=row, column=1)

total_of_circuits_label = ttk.Label(frame_1, text='Total of sub-networks:')
total_of_circuits_label.grid(row=row, column=2)
total_of_circuits_label = ttk.Label(frame_1, textvariable=counter_of_sub_networks)
total_of_circuits_label.grid(row=row, column=3)

row = row + 1

type_of_interconnection_label = ttk.Label(frame_1, text='Type of interconnection')
type_of_interconnection_label.grid(row=row, column=0)
type_of_interconnection_combobox = ttk.Combobox(frame_1, values=cfg.INTERCONNECTION_TYPES, state='normal', textvariable=reset_type_of_connection_combobox)
type_of_interconnection_combobox.grid(row=row, column=1)
type_of_interconnection_combobox.current()

row = row + 1

spacer_1.grid(row=row, column=0)

row = row + 1

element_A_label = ttk.Label(frame_1, text='Element A')
element_A_label.grid(row=row, column=0)
element_A_combobox = ttk.Combobox(frame_1, values=cfg.ELEMENT_TYPES, state='normal', textvariable=reset_element_A_combobox)
element_A_combobox.grid(row=row, column=1)
element_A_value_label = ttk.Label(frame_1, text='Value')
element_A_value_label.grid(row=row, column=2)
element_A_entry = ttk.Entry(frame_1, state='normal', textvariable=reset_element_A_entry)
element_A_entry.grid(row=row, column=3)

row = row + 1

element_B_label = ttk.Label(frame_1, text='Element B')
element_B_label.grid(row=row, column=0)
element_B_combobox = ttk.Combobox(frame_1, values=cfg.ELEMENT_TYPES, state='normal', textvariable=reset_element_B_combobox)
element_B_combobox.grid(row=row, column=1)
element_B_value_label = ttk.Label(frame_1, text='Value')
element_B_value_label.grid(row=row, column=2)
element_B_entry = ttk.Entry(frame_1, state='normal', textvariable=reset_element_B_entry)
element_B_entry.grid(row=row, column=3)

row = row + 1

element_C_label = ttk.Label(frame_1, text='Element C')
element_C_label.grid(row=row, column=0)
element_C_combobox = ttk.Combobox(frame_1, values=cfg.ELEMENT_TYPES, state='normal', textvariable=reset_element_C_combobox)
element_C_combobox.grid(row=row, column=1)
element_C_value_label = ttk.Label(frame_1, text='Value')
element_C_value_label.grid(row=row, column=2)
element_C_entry = ttk.Entry(frame_1, state='normal', textvariable=reset_element_C_entry)
element_C_entry.grid(row=row, column=3)

row = row + 1

spacer_2.grid(row=row, column=0)

row = row + 1

touchstone_file_button = ttk.Button(frame_1, text='Add touchstone file', command=add_touchstone_file)
touchstone_file_button.grid(row=row, column=0)
touchstone_file_entry = ttk.Entry(frame_1, textvariable=reset_touchstone_file_entry)
touchstone_file_entry.grid(row=row, column=1)

row = row + 1

spacer_3.grid(row=row, column=0)

row = row + 1

add_sub_network_button = ttk.Button(frame_1, text='Add sub-network', state='normal', command=add_sub_network)
add_sub_network_button.grid(row=row, column=2)

row = row + 1

save_sub_networks_button = ttk.Button(frame_1, text='Save sub-networks', state='disable', command=save_sub_networks)
save_sub_networks_button.grid(row=row, column=2)

row = row + 1

spacer_4.grid(row=row, column=0)

# Start of frame 2 #
row = row + 1

frame_2 = ttk.Frame(mainWindow)
frame_2.grid(row=row, column=0)

spacer_5 = ttk.Label(frame_2)
spacer_6 = ttk.Label(frame_2)

row = row + 1

start_frequency_label = ttk.Label(frame_2, text = 'Start frequency')
start_frequency_label.grid(row=row, column=0)
start_frequency_entry = ttk.Entry(frame_2, state='disable', textvariable=reset_start_frequency_entry)
start_frequency_entry.grid(row=row, column=1)
start_frequency_combobox = ttk.Combobox(frame_2, state='disable', values = cfg.FREQUENCY_PREFIXES)
start_frequency_combobox.grid(row=row, column=2)

row = row + 1

end_frequency_label = ttk.Label(frame_2, text='End frequency')
end_frequency_label.grid(row=row, column=0)
end_frequency_entry = ttk.Entry(frame_2, state='disable', textvariable=reset_end_frequency_entry)
end_frequency_entry.grid(row=row, column=1)
end_frequency_combobox = ttk.Combobox(frame_2, state='disable', values=cfg.FREQUENCY_PREFIXES)
end_frequency_combobox.grid(row=row, column=2)

row = row + 1

analysis_frequency_step_label = ttk.Label(frame_2, text='Analysis frequency step')
analysis_frequency_step_label.grid(row=row, column=0)
analysis_frequency_step_entry = ttk.Entry(frame_2, state='disable', textvariable=reset_analysis_frequency_step_entry)
analysis_frequency_step_entry.grid(row=row, column=1)
analysis_frequency_step_combobox = ttk.Combobox(frame_2, state='disable', values = cfg.FREQUENCY_PREFIXES)
analysis_frequency_step_combobox.grid(row=row, column=2)

row = row + 1

spacer_5.grid(row=row, column=0)

row = row + 1

save_frequency_parameters_button = ttk.Button(frame_2, text='Save frequency parameters', state='disable', command=save_frequency_parameters)
save_frequency_parameters_button.grid(row=row, column=0)

row = row + 1

spacer_6.grid(row=row, column=0)

row = row + 1

calculate_parameters_button = ttk.Button(frame_2, text='Calculate parameters', state='disable', command=calculate_parameters)
calculate_parameters_button.grid(row=row, column=0)
parameters_to_calculate_combobox = ttk.Combobox(frame_2, values=cfg.NETWORK_PARAMETERS, state='disable')
parameters_to_calculate_combobox.grid(row=row, column=1)

row = row + 1

plot_parameters_button = ttk.Button(frame_2, text='Plot parameters', state='disable', command=plot_parameters)
plot_parameters_button.grid(row=row, column=0)
plot_parameters_in_format_combobox = ttk.Combobox(frame_2, values=cfg.PLOT_FORMATS, state='disable')
plot_parameters_in_format_combobox.grid(row=row, column=1)

mainWindow.mainloop()