import tkinter as tk
from tkinter import ttk

from openpyxl import *
import config_parameters as cfg

excel_workbook = load_workbook(filename = cfg.EXCEL_BASE_TEMPLATE_FILE)
excel_network_info_sheet = excel_workbook[cfg.EXCEL_NETWORK_INFO_SHEET]
excel_network_parameters_sheet = excel_workbook[cfg.EXCEL_NETWORK_PARAMETERS_SHEET]
excel_circuit_counter_row = cfg.EXCEL_INITIAL_ROW_NETWORK_ID

def save_frecuency_parameters():
	#Create excel

	start_frecuency = start_frecuency_entry.get()
	start_frecuency_prefix = start_frecuency_combobox.get()
	end_frecuency = end_frecuency_entry.get()
	end_frecuency_prefix = end_frecuency_combobox.get()
	frecuency_points = frecuency_points_entry.get()

	if start_frecuency == '' or start_frecuency_prefix == '' or end_frecuency == '' or end_frecuency_prefix == '' or frecuency_points == '':
		pass
	else:
		excel_network_info_sheet.cell(row = cfg.EXCEL_INITIAL_ROW, column = cfg.EXCEL_COLUMN_B, value = start_frecuency + start_frecuency_prefix)
		excel_network_info_sheet.cell(row = cfg.EXCEL_INITIAL_ROW + 1, column = cfg.EXCEL_COLUMN_B, value = end_frecuency + end_frecuency_prefix)
		excel_network_info_sheet.cell(row = cfg.EXCEL_INITIAL_ROW + 2, column = cfg.EXCEL_COLUMN_B, value = frecuency_points)
		excel_workbook.save(cfg.EXCEL_NETWORK_PARAMETERS_FILE)
		start_frecuency_entry.config(state = 'disable')
		start_frecuency_combobox.config(state = 'disable')
		end_frecuency_entry.config(state = 'disable')
		end_frecuency_combobox.config(state = 'disable')
		frecuency_points_entry.config(state = 'disable')
		type_of_network_combobox.config(state = 'normal')
		parameters_to_calculate_combobox.config(state = 'normal')
		save_network_parameters_button.config(state = 'normal')


def save_network_parameters():
	type_of_network = type_of_network_combobox.get()
	parameters_to_calculate = parameters_to_calculate_combobox.get()

	if type_of_network == '' and parameters_to_calculate == '':
		pass
	else:
		excel_network_info_sheet.cell(row = cfg.EXCEL_INITIAL_ROW + 4, column = cfg.EXCEL_COLUMN_B, value = type_of_network)
		excel_network_info_sheet.cell(row = cfg.EXCEL_INITIAL_ROW + 5, column = cfg.EXCEL_COLUMN_B, value = parameters_to_calculate)
		excel_workbook.save(cfg.EXCEL_NETWORK_PARAMETERS_FILE)
		type_of_network_combobox.config(state = 'disable')
		parameters_to_calculate_combobox.config(state = 'disable')
		type_of_circuit_combobox.config(state = 'normal')
		type_of_connection_combobox.config(state = 'normal')
		element_A_combobox.config(state = 'normal')
		element_A_entry.config(state = 'normal')
		element_B_combobox.config(state = 'normal')
		element_B_entry.config(state = 'normal')
		element_C_combobox.config(state = 'normal')
		element_C_entry.config(state = 'normal')
		add_circuit_button.config(state = 'normal')


def add_sub_network():
	type_of_circuit = type_of_circuit_combobox.get()
	type_of_connection = type_of_connection_combobox.get()
	element_A = element_A_combobox.get()
	element_A_value = element_A_entry.get()
	element_B = element_B_combobox.get()
	element_B_value = element_B_entry.get()
	element_C = element_C_combobox.get()
	element_C_value = element_C_entry.get()

	if 0:
		pass
	else:
		global excel_circuit_counter_row

		counter = counter_of_sub_networks.get()
		counter = counter + 1

		if type_of_connection == '':
			type_of_connection = 'Cascade'

		excel_network_info_sheet.cell(row=excel_circuit_counter_row, column=cfg.EXCEL_COLUMN_A, value='Network ' + str(counter))
		excel_network_info_sheet.cell(row = excel_circuit_counter_row, column = cfg.EXCEL_COLUMN_B, value = type_of_connection)
		excel_network_info_sheet.cell(row = excel_circuit_counter_row, column = cfg.EXCEL_COLUMN_C, value = type_of_circuit)
		excel_network_info_sheet.cell(row = excel_circuit_counter_row, column = cfg.EXCEL_COLUMN_D, value = element_A)
		excel_network_info_sheet.cell(row = excel_circuit_counter_row, column = cfg.EXCEL_COLUMN_E, value = element_A_value)
		excel_network_info_sheet.cell(row = excel_circuit_counter_row, column = cfg.EXCEL_COLUMN_F, value = element_B)
		excel_network_info_sheet.cell(row = excel_circuit_counter_row, column = cfg.EXCEL_COLUMN_G, value = element_B_value)
		excel_network_info_sheet.cell(row = excel_circuit_counter_row, column = cfg.EXCEL_COLUMN_H, value = element_C)
		excel_network_info_sheet.cell(row = excel_circuit_counter_row, column = cfg.EXCEL_COLUMN_I, value = element_C_value)
		excel_workbook.save(cfg.EXCEL_NETWORK_PARAMETERS_FILE)

		excel_circuit_counter_row = excel_circuit_counter_row + 1

		#counter = counter_of_sub_networks.get()
		#counter = counter + 1

		counter_of_sub_networks.set(counter)
		reset_1.set('')
		reset_2.set('')
		reset_3.set('')
		reset_4.set('')
		reset_5.set('')
		reset_6.set('')
		reset_7.set('')
		reset_8.set('')


def calculate_parameters():
	pass

mainWindow = tk.Tk()
mainWindow.title("Two-port network parameter calculator")
mainWindow.geometry('800x600')

counter_of_sub_networks = tk.IntVar(mainWindow, 0)
reset_1 = tk.StringVar(mainWindow, '')
reset_2 = tk.StringVar(mainWindow, '')
reset_3 = tk.StringVar(mainWindow, '')
reset_4 = tk.StringVar(mainWindow, '')
reset_5 = tk.StringVar(mainWindow, '')
reset_6 = tk.StringVar(mainWindow, '')
reset_7 = tk.StringVar(mainWindow, '')
reset_8 = tk.StringVar(mainWindow, '')

row = 0

start_frecuency_label = ttk.Label(mainWindow, text = 'Start frecuency')
start_frecuency_label.grid(row = row, column = 0)
start_frecuency_entry = ttk.Entry(mainWindow)
start_frecuency_entry.grid(row = row, column = 1)
start_frecuency_combobox = ttk.Combobox(mainWindow, values = cfg.FRECUENCY_PREFIXES)
start_frecuency_combobox.grid(row = row, column = 2)

row = row + 1

end_frecuency_label = ttk.Label(mainWindow, text='End frecuency')
end_frecuency_label.grid(row=row, column=0)
end_frecuency_entry = ttk.Entry(mainWindow)
end_frecuency_entry.grid(row=row, column=1)
end_frecuency_combobox = ttk.Combobox(mainWindow, values=cfg.FRECUENCY_PREFIXES)
end_frecuency_combobox.grid(row=row, column=2)

row = row + 1

frecuency_points_label = ttk.Label(mainWindow, text='Frecuency points')
frecuency_points_label.grid(row=row, column=0)
frecuency_points_entry = ttk.Entry(mainWindow)
frecuency_points_entry.grid(row=row, column=1)

row = row + 1

spacer_1 = ttk.Label(mainWindow)
spacer_1.grid(row=row, column=0)

row = row + 1

save_frecuency_parameters_button = ttk.Button(mainWindow, text='Save frecuency parameters',
											  command=save_frecuency_parameters)
save_frecuency_parameters_button.grid(row=row, column=2)

row = row + 1

spacer_1 = ttk.Label(mainWindow)
spacer_1.grid(row=row, column=0)

row = row + 1

type_of_netwwork_label = ttk.Label(mainWindow, text='Type of network')
type_of_netwwork_label.grid(row=row, column=0)
type_of_network_combobox = ttk.Combobox(mainWindow, values=cfg.NETWORK_TYPES, state='disable')
type_of_network_combobox.grid(row=row, column=1)

row = row + 1

spacer_3 = ttk.Label(mainWindow)
spacer_3.grid(row=row, column=0)

row = row + 1

parameters_to_calculate_label = ttk.Label(mainWindow, text='Parameters to calculate')
parameters_to_calculate_label.grid(row=row, column=0)
parameters_to_calculate_combobox = ttk.Combobox(mainWindow, values=cfg.NETWORK_PARAMETERS, state='disable')
parameters_to_calculate_combobox.grid(row=row, column=1)

#In case of N-port uncomment the next lines
#total_of_ports_label = ttk.Label(mainWindow, text='Total of ports')
#total_of_ports_label.grid(row=row, column=2)
#total_of_ports_entry = ttk.Entry(mainWindow)
#total_of_ports_entry.grid(row=row, column=3)
#total_of_ports_entry.config(state='normal')

row = row + 1

spacer_1 = ttk.Label(mainWindow)
spacer_1.grid(row=row, column=0)

row = row + 1

save_network_parameters_button = ttk.Button(mainWindow, text='Save network parameters', state='disable',
											command=save_network_parameters)
save_network_parameters_button.grid(row=row, column=2)

row = row + 1

spacer_4 = ttk.Label(mainWindow)
spacer_4.grid(row=row, column=0)

row = row + 1

type_of_circuit_label = ttk.Label(mainWindow, text='Type of circuit')
type_of_circuit_label.grid(row=row, column=0)
type_of_circuit_combobox = ttk.Combobox(mainWindow, values=cfg.CIRCUIT_TYPES, state='disable',
										textvariable=reset_1)
type_of_circuit_combobox.grid(row=row, column=1)

total_of_circuits_label = ttk.Label(mainWindow, text='Total of sub-networks:')
total_of_circuits_label.grid(row=row, column=2)
total_of_circuits_label = ttk.Label(mainWindow, textvariable=counter_of_sub_networks)
total_of_circuits_label.grid(row=row, column=3)

row = row + 1

type_of_connection_label = ttk.Label(mainWindow, text='Type of connection')
type_of_connection_label.grid(row=row, column=0)
type_of_connection_combobox = ttk.Combobox(mainWindow, values=cfg.CONNECTION_TYPES, state='disable',
										   textvariable=reset_2)
type_of_connection_combobox.grid(row=row, column=1)

row = row + 1

spacer_5 = ttk.Label(mainWindow)
spacer_5.grid(row=row, column=0)

row = row + 1

element_A_label = ttk.Label(mainWindow, text='Element A')
element_A_label.grid(row=row, column=0)
element_A_combobox = ttk.Combobox(mainWindow, values=cfg.ELEMENT_TYPES, state='disable', textvariable=reset_3)
element_A_combobox.grid(row=row, column=1)
element_A_value_label = ttk.Label(mainWindow, text='Value')
element_A_value_label.grid(row=row, column=2)
element_A_entry = ttk.Entry(mainWindow, state='disable', textvariable=reset_4)
element_A_entry.grid(row=row, column=3)

row = row + 1

element_B_label = ttk.Label(mainWindow, text='Element B')
element_B_label.grid(row=row, column=0)
element_B_combobox = ttk.Combobox(mainWindow, values=cfg.ELEMENT_TYPES, state='disable', textvariable=reset_5)
element_B_combobox.grid(row=row, column=1)
element_B_value_label = ttk.Label(mainWindow, text='Value')
element_B_value_label.grid(row=row, column=2)
element_B_entry = ttk.Entry(mainWindow, state='disable', textvariable=reset_6)
element_B_entry.grid(row=row, column=3)

row = row + 1

element_C_label = ttk.Label(mainWindow, text='Element C')
element_C_label.grid(row=row, column=0)
element_C_combobox = ttk.Combobox(mainWindow, values=cfg.ELEMENT_TYPES, state='disable', textvariable=reset_7)
element_C_combobox.grid(row=row, column=1)
element_C_value_label = ttk.Label(mainWindow, text='Value')
element_C_value_label.grid(row=row, column=2)
element_C_entry = ttk.Entry(mainWindow, state='disable', textvariable=reset_8)
element_C_entry.grid(row=row, column=3)

row = row + 1

add_circuit_button = ttk.Button(mainWindow, text='Add sub-network', state='disable', command=add_sub_network)
add_circuit_button.grid(row=row, column=2)
# add_circuit_button = ttk.Button(mainWindow, text='Add circuit', state='disable', command=add_sub_network)
# add_circuit_button.grid(row=row, column=3)

row = row + 1

calculate_parameters_button = ttk.Button(mainWindow, text='Calculate parameters', state='normal')
calculate_parameters_button.grid(row=row, column=2)

mainWindow.mainloop()
